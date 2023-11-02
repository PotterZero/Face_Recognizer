import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import openpyxl
#import Webcam

# Thư mục chứa hình ảnh khuôn mặt
path = "dataset"
images = []
classNames = []
myList = os.listdir(path)

# Load danh sách tên từ tệp Excel
current_date = datetime.now().strftime("%Y-%m-%d")
filename = f"{current_date}_cham_cong.xlsx"
try:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    nameList = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách khuôn mặt"
    sheet.cell(row=1, column=1, value="Tên")
    sheet.cell(row=1, column=2, value="Chấm công")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    nameList = []

for cl in myList:
    curImg = cv2.imread(os.path.join(path, cl))
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
# Encode danh sách khuôn mặt
def Mahoa(images): #images khai báo rỗng bên trên
    encodeList = []
    for img in images:
        face_encodings = face_recognition.face_encodings(img) # sử dụng thư viện mã hóa khuôn mặt trả về giá trị
        if face_encodings:
            encode = face_encodings[0] #lấy 1 tên để k trùng
            encodeList.append(encode) #add encode
    return encodeList

encodeListKnown = Mahoa(images)
print("Encoding complete")

# Hàm thêm chấm công vào tệp Excel
def thamgia(name):
    if name not in nameList:
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')
        sheet.append([name, dtString])
        workbook.save(filename)
        nameList.append(name)

# Khởi động webcam và xử lý hình ảnh
cap = cv2.VideoCapture(0)
cap.set(3, 320)
cap.set(4, 180)

minW = 0.1 *cap.get(3)
minH = 0.1 *cap.get(4)

if not cap.isOpened():
    print("Cannot open camera")
    exit()

while True:
    ret, frame = cap.read()
    facecurFrame = face_recognition.face_locations(frame)
    face_encodings = face_recognition.face_encodings(frame, facecurFrame)

    # Chỉ phát hiện khuôn mặt trong vùng hình ảnh nhỏ
    for encodeFace, faceLoc in zip(face_encodings, facecurFrame):
        if faceLoc[2] - faceLoc[0] > minW and faceLoc[3] - faceLoc[1] > minH:
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex]
                thamgia(name)

    cv2.imshow('Cam', frame)

    if cv2.waitKey(1) == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
