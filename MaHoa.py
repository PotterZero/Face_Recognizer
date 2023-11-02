import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import openpyxl
#import Webcam

# Thư mục chứa hình ảnh khuôn mặt
path = "dataset"
images = []
classNames = []
myList = os.listdir(path)

# Load danh sách tên từ tệp Excel
current_date = datetime.now().strftime("%Y-%m-%d")
filename = f"{current_date}_cham_cong.xlsx"
try:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    nameList = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách khuôn mặt"
    sheet.cell(row=1, column=1, value="Tên")
    sheet.cell(row=1, column=2, value="Chấm công")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    nameList = []

for cl in myList:
    curImg = cv2.imread(os.path.join(path, cl))
    images.append(curImg) #đọc
    classNames.append(os.path.splitext(cl)[0])
# Encode danh sách khuôn mặt
def Mahoa(images): #images khai báo rỗng bên trên
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_encodings = face_recognition.face_encodings(img) # sử dụng thư viện mã hóa khuôn mặt trả về giá trị
        if face_encodings:
            encode = face_encodings[0] #lấy 1 tên để k trùng
            encodeList.append(encode) #add encode
    return encodeList

encodeListKnown = Mahoa(images)
print("Encoding complete")

# Hàm thêm chấm công vào tệp Excel
def thamgia(name):
    if name not in nameList:
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')
        sheet.append([name, dtString])
        workbook.save(filename)
        nameList.append(name)

# Khởi động webcam và xử lý hình ảnh
cap = cv2.VideoCapture(0)
cap.set(3, 320)
cap.set(4, 180)

minW = 0.1 *cap.get(3)
minH = 0.1 *cap.get(4)

if not cap.isOpened():
    print("Cannot open camera")
    exit()

while True:
    ret, frame = cap.read()
    framS = cv2.resize(frame, (0, 0), None, fx=0.5, fy=0.5)
    framS = cv2.cvtColor(framS, cv2.COLOR_BGR2RGB)
    facecurFrame = face_recognition.face_locations(framS)
    encodecurFrame = face_recognition.face_encodings(framS)

    for encodeFace, faceLoc in zip(encodecurFrame, facecurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if faceDis[matchIndex] < 0.50:
            name = classNames[matchIndex].upper()
            thamgia(name)
        else:
            name = 'Unknown'
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = y1 * 2, x2 * 2, y2 * 2, x1 * 2
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(frame, name, (x2, y2), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

    cv2.imshow('Cam', frame)
    if cv2.waitKey(1) == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()